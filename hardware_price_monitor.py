#!/usr/bin/env python3
"""
MVP - Monitor de Precios de Hardware
======================================
Monitorea el precio de 4 componentes clave frente a 2 competidores.
Ejecucion manual o programada (cron / tarea programada).

Input:
    mapeo_hardware.xlsx  - Excel con columnas:
        SKU_Interno, Nombre_Producto, Precio_Propio, Stock_Propio,
        URL_Competidor_A, URL_Competidor_B

Output:
    Dashboard_Hardware_YYYYMMDD.xlsx
"""

import os
import sys
import ssl
import shutil
import smtplib
import argparse
import logging
from datetime import datetime
from email.message import EmailMessage

import pandas as pd
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv

BASE_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "es-ES,es;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )


def parse_price(text: str | None) -> float | None:
    if not text:
        return None
    cleaned = text.replace("\u20ac", "").replace("\xa0", " ").strip()
    cleaned = cleaned.replace(".", "")
    cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def scrape_competitor(url: str, label: str, sku: str) -> float | None:
    try:
        resp = requests.get(url, headers=BASE_HEADERS, timeout=15)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        price_text = None
        selectors = [
            ".price",
            ".product-price",
            ".precio",
            '[data-price]',
            "span.price",
            "div.price",
            "span.woocommerce-Price-amount",
            ".current-price",
            ".our_price",
            "#price",
            '[itemprop="price"]',
            '[data-product-price]',
            'span[class*="price"]',
            'div[class*="price"]',
            ".precio-final",
            ".precio_actual",
            ".price-value",
        ]
        for sel in selectors:
            el = soup.select_one(sel)
            if el:
                attrs = ("content", "data-price", "data-product-price")
                for attr in attrs:
                    val = el.get(attr)
                    if val:
                        price_text = val
                        break
                if not price_text:
                    price_text = el.get_text(strip=True)
                if price_text:
                    break

        if not price_text:
            for tag in ("span", "div", "p", "strong", "ins", "bdi"):
                for el in soup.find_all(tag):
                    txt = el.get_text(strip=True)
                    if "\u20ac" in txt:
                        price_text = txt
                        break
                if price_text:
                    break

        if price_text:
            return parse_price(price_text)

        logging.warning("No se encontro precio en el HTML para SKU %s - %s", sku, label)
        return None
    except Exception as exc:
        logging.warning("Error en SKU %s - %s: %s", sku, label, exc)
        return None


def send_email(dashboard_file: str, summary: str) -> None:
    sender = os.getenv("GMAIL_SENDER")
    password = os.getenv("GMAIL_APP_PASSWORD")
    recipient = os.getenv("GMAIL_RECIPIENT")

    if not all([sender, password, recipient]):
        logging.warning("Faltan variables SMTP en .env - correo no enviado")
        return

    msg = EmailMessage()
    msg["Subject"] = f"Dashboard Hardware - {datetime.now().strftime('%Y-%m-%d')}"
    msg["From"] = sender
    msg["To"] = recipient
    msg.set_content(summary)

    try:
        with open(dashboard_file, "rb") as f:
            data = f.read()
        msg.add_attachment(
            data,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(dashboard_file),
        )
    except Exception as exc:
        logging.warning("No se pudo adjuntar el dashboard: %s", exc)

    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender, password)
            server.send_message(msg)
        logging.info("Correo enviado exitosamente a %s", recipient)
    except Exception as exc:
        logging.error("Error al enviar correo: %s", exc)


def main() -> None:
    setup_logging()
    load_dotenv()

    parser = argparse.ArgumentParser(
        description="Monitor de Precios de Hardware",
        epilog=(
            "Escanea los precios de productos propios frente a 2 competidores "
            "a partir de un archivo Excel, genera un dashboard con alertas y "
            "envia el reporte por correo electronico via SMTP de Google.\n\n"
            "Ejemplo:\n"
            "  python3 hardware_price_monitor.py -f mapeo_hardware.xlsx"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        add_help=False,
    )
    parser.add_argument(
        "-help", "-h", "--help",
        action="help",
        help="Muestra este mensaje de ayuda y sale",
    )
    parser.add_argument(
        "-f", "--file",
        default="mapeo_hardware.xlsx",
        help="Archivo Excel de entrada (default: mapeo_hardware.xlsx)",
    )
    args = parser.parse_args()

    input_file = args.file
    if not os.path.isfile(input_file):
        logging.error("No se encuentra el archivo de entrada: %s", input_file)
        sys.exit(1)

    df = pd.read_excel(input_file)
    required_cols = {
        "SKU_Interno", "Nombre_Producto", "Precio_Propio", "Stock_Propio",
        "URL_Competidor_A", "URL_Competidor_B",
    }
    missing = required_cols - set(df.columns)
    if missing:
        logging.error("Columnas faltantes en el Excel: %s", missing)
        sys.exit(1)

    results = []
    scrape_errors = 0
    total_skus = len(df)

    logging.info("Iniciando monitoreo de %d producto(s)...", total_skus)

    for _, row in df.iterrows():
        sku = row["SKU_Interno"]
        name = row["Nombre_Producto"]
        url_a = row["URL_Competidor_A"]
        url_b = row["URL_Competidor_B"]

        logging.info("[%s] Procesando %s...", sku, name)

        # 1) Datos locales desde el Excel
        our_price = row.get("Precio_Propio")
        stock = row.get("Stock_Propio")

        # 2) Scraping competidores
        price_a = scrape_competitor(url_a, "Competidor A", sku) if pd.notna(url_a) else None
        price_b = scrape_competitor(url_b, "Competidor B", sku) if pd.notna(url_b) else None

        if price_a is None and pd.notna(url_a):
            scrape_errors += 1
            logging.error("No se pudo obtener el precio de %s", url_a)
        if price_b is None and pd.notna(url_b):
            scrape_errors += 1
            logging.error("No se pudo obtener el precio de %s", url_b)

        # 3) Calculos
        comp_prices = [p for p in (price_a, price_b) if p is not None]
        min_comp = min(comp_prices) if comp_prices else None

        diff = None
        if our_price is not None and min_comp is not None:
            diff = round(our_price - min_comp, 2)

        status = "ALERTA" if diff is not None and diff > 0 else "OK"

        results.append(
            {
                "SKU_Interno": sku,
                "Nombre_Producto": name,
                "Precio_Propio": our_price,
                "Stock_Propio": stock,
                "Precio_Comp_A": price_a,
                "Precio_Comp_B": price_b,
                "Precio_Min_Competencia": min_comp,
                "Diferencia": diff,
                "Status": status,
            }
        )

        logging.info(
            "  -> Propio: %s | Stock: %s | CompA: %s | CompB: %s | MinComp: %s | Diff: %s | %s",
            our_price, stock, price_a, price_b, min_comp, diff, status,
        )

    # 4) Exportar dashboard
    out_df = pd.DataFrame(results)
    today = datetime.now().strftime("%Y%m%d")
    out_file = f"Dashboard_Hardware_{today}.xlsx"

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        out_df.to_excel(writer, index=False, sheet_name="Dashboard")
        ws = writer.sheets["Dashboard"]

        from openpyxl.styles import PatternFill, Font

        alert_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        alert_font = Font(bold=True, color="CC0000")
        status_col = list(out_df.columns).index("Status") + 1

        for row_idx, status_val in enumerate(out_df["Status"], start=2):
            if status_val == "ALERTA":
                ws.cell(row=row_idx, column=status_col).fill = alert_fill
                ws.cell(row=row_idx, column=status_col).font = alert_font

    # 5) Resumen para el correo
    alerts = [r for r in results if r["Status"] == "ALERTA"]
    summary_lines = [
        "=== Resumen del Monitoreo de Precios ===",
        f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Total SKUs procesados: {total_skus}",
        f"Alertas de precio: {len(alerts)}",
        f"Errores de scraping: {scrape_errors}",
        "",
    ]
    if alerts:
        summary_lines.append("Productos con ALERTA:")
        for a in alerts:
            summary_lines.append(
                f"  - {a['Nombre_Producto']} (SKU: {a['SKU_Interno']}) "
                f"| Propio: {a['Precio_Propio']} | "
                f"Mín Comp: {a['Precio_Min_Competencia']} | "
                f"Diferencia: {a['Diferencia']}"
            )
    else:
        summary_lines.append("No hay alertas de precio.")

    summary_lines.append("")
    summary_lines.append(f"Dashboard adjunto: {out_file}")
    summary = "\n".join(summary_lines)

    logging.info("=" * 50)
    logging.info("PROCESO COMPLETADO")
    logging.info("  %d/%d SKUs procesados", total_skus, total_skus)
    logging.info("  %d alerta(s) de precio", len(alerts))
    logging.info("  %d error(es) de scraping ignorado(s)", scrape_errors)
    logging.info("  Dashboard exportado: %s", out_file)

    # 6) Enviar correo
    send_email(out_file, summary)

    # 7) Copiar al escritorio
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    if os.path.isdir(desktop):
        try:
            shutil.copy(out_file, desktop)
            logging.info("  Dashboard copiado a: %s", desktop)
        except Exception as exc:
            logging.warning("No se pudo copiar al escritorio: %s", exc)


if __name__ == "__main__":
    main()
